#!/usr/bin/env python3

import json
import re
import os
import glob
import socket
import argparse
import sys
import create_excel
from openpyxl.styles import Font
import openpyxl

description = """ This script will parse files"""
epilog = '''Report bugs to johan.vikerborn@gmail.com.'''

# Get the current working directory
cwd = os.getcwd()
os.chdir(cwd)

# Make a List of all the Json files in the cwd
json_files = glob.glob('*.json')


def run_parser(json_files, filename='output.xlsx'):
    if os.path.isfile(filename):
        wb = openpyxl.load_workbook(filename)
        try:
            sheet = wb.get_sheet_by_name('Switch_info')
            if sheet.cell(row=1, column=1).value == None:
                create_excel.createExcel(filename)
        except:
            print("There Is Something Wrong With The File, Please Remove It and Try Again")
    else:
        create_excel.createExcel(filename)

    # Get the hostnames of the Switches from filenames
    pattern = re.compile(r'ip_interface_facts-([^,\s]+).json')
    matches = filter(pattern.search, json_files)

    interface_hostnames = []
    if matches:
        for match in matches:
            interface_hostnames.append(pattern.search(match).group(1))

    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    # ws = wb.active
    # targetFileActiveSheet = filename.get_sheet_by_name(sheet)

    # Run parser code
    counter = 0

    item_list_interface = {}
    item_list_vrf = {}

    for item in interface_hostnames:
        #
        with open('ip_interface_facts-{}.json'.format(interface_hostnames[counter])) as interface_file, open('ip_vrf_interface_facts-{}.json'.format(interface_hostnames[counter])) as vrf_file:

            data = json.loads(interface_file.read())
            data_vrf = json.loads(vrf_file.read())

            keydata = data.keys()
            # keydata2 = data_vrf.keys()
            ip_list = ['10\.2\.', '10\.4\.', '10\.12\.']
            for i in data:

                # Use a variable to collect the IP address for every interface and then compare it to the ip list.
                jsonIpAddress = data[i]['config']['ip_address']

                for pattern in ip_list:
                    x = re.search(pattern, str(jsonIpAddress))
                    if (x):

                        item_list_interface.update({data[i]['config']['name']: {'name': data[i]['config']['name'], 'hostname': interface_hostnames[counter], 'ipaddress': data[i]['config']['ip_address'], 'mtu': data[i]['config']['mtu'], 'helper': data[i]
                                                                                ['config']['helper'],  'inbound_acl': data[i]['config']['inbound_acl'], 'outgoing_acl': data[i]['config']['outgoing_acl'], 'proxy_arp': data[i]['config']['proxy_arp'], 'unicast_rpf': data[i]['config']['unicast_rpf']}})

            counter += 1

            ip_list = ['10\.2\.', '10\.4\.', '10\.12\.']
            for i in data_vrf:
                jsonIpAddress = data_vrf[i]['data']['ip_address']
                # print(i)

                for pattern in ip_list:
                    x = re.search(pattern, str(jsonIpAddress))

                    if (x):
                        item_list_vrf.update({data_vrf[i]['data']['name']: {
                                             'vrf': data_vrf[i]['data']['vrf'], 'name': data_vrf[i]['data']['name'], 'protocol_state': data_vrf[i]['data']['protocol_state']}})

    for k, v in item_list_interface.items():
        try:
            excel_data = [(item_list_interface[k]['hostname'], item_list_interface[k]['name'], item_list_interface[k]['ipaddress'], item_list_vrf[k]['vrf'], item_list_vrf[k]['name'], item_list_interface[k]
                           ['mtu'], item_list_interface[k]['helper'], item_list_interface[k]['inbound_acl'], item_list_interface[k]['outgoing_acl'], item_list_interface[k]['proxy_arp'], item_list_interface[k]['unicast_rpf'])]
            for i in excel_data:
                sheet.append(i)
                wb.save(filename)
        except KeyError:
            excel_data = [(item_list_interface[k]['hostname'], item_list_interface[k]['name'], item_list_interface[k]['ipaddress'], ' ', ' ', item_list_interface[k]
                           ['mtu'], item_list_interface[k]['helper'], item_list_interface[k]['inbound_acl'], item_list_interface[k]['outgoing_acl'], item_list_interface[k]['proxy_arp'], item_list_interface[k]['unicast_rpf'])]
            for i in excel_data:
                sheet.append(i)
                wb.save(filename)


def createExcel(filename):
    create_excel.createExcel(filename)


# Argparse: Make a cmd flags to script and help information
parser = argparse.ArgumentParser(description=description, epilog=epilog)
parser.add_argument('-e', '--excelfile', help='enter namn of existing excel file')
parser.add_argument('-E', '--create', help='Create a new excel file with headers')
args = parser.parse_args()

if args.excelfile:
    run_parser(filename=args.excelfile)
elif args.create:
    createExcel(filename=args.create)
else:
    run_parser(json_files)
